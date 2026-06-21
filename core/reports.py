"""
core/reports.py — Swahilipot Hub Premium Report Engine
Professional PDF (ReportLab) and Excel (openpyxl) generation.
"""
from io import BytesIO
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    Workbook = None

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable, KeepTogether,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
except ImportError:
    rl_canvas = None

from django.http import HttpResponse
from django.contrib.staticfiles import finders

# ── Brand palette ──────────────────────────────────────────────────────────
BRAND_NAME     = "Swahilipot Hub Foundation"
BRAND_TAGLINE  = "Empowering Youth Through Technology, Arts & Entrepreneurship"
BRAND_ADDRESS  = "Mombasa County Governor's Office, Mombasa, Kenya"
BRAND_WEBSITE  = "www.swahilipothub.co.ke"
BRAND_EMAIL    = "info@swahilipothub.co.ke"
BRAND_PHONE    = "+254 114 635 505"
BRAND_LOGO_STATIC_PATH = "images/swahilipot-logo.jfif"

# Hex values (shared by both engines)
HEX_NAVY    = "0F2B5B"   # deep navy
HEX_BLUE    = "1E40AF"   # brand blue
HEX_CYAN    = "0EA5E9"   # accent cyan
HEX_GOLD    = "F59E0B"   # brand gold
HEX_LIGHT   = "EFF6FF"   # row alt background
HEX_STRIPE  = "DBEAFE"   # header light stripe
HEX_DARK    = "111827"   # body text
HEX_MID     = "6B7280"   # secondary text
HEX_BORDER  = "93C5FD"   # subtle border

if rl_canvas:
    C_NAVY   = colors.HexColor(f"#{HEX_NAVY}")
    C_BLUE   = colors.HexColor(f"#{HEX_BLUE}")
    C_CYAN   = colors.HexColor(f"#{HEX_CYAN}")
    C_GOLD   = colors.HexColor(f"#{HEX_GOLD}")
    C_LIGHT  = colors.HexColor(f"#{HEX_LIGHT}")
    C_WHITE  = colors.white
    C_DARK   = colors.HexColor(f"#{HEX_DARK}")
    C_MID    = colors.HexColor(f"#{HEX_MID}")


def _logo_path():
    return finders.find(BRAND_LOGO_STATIC_PATH)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def excel_response(filename, headers, rows, subtitle=""):
    if Workbook is None:
        # Fallback to CSV
        text = "\n".join(
            [",".join(str(h) for h in headers)] +
            [",".join('"' + str(v).replace('"', '""') + '"' for v in row) for row in rows]
        )
        resp = HttpResponse(text, content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        return resp

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.sheet_view.showGridLines = False

    # ── colour palette ─────────────────────────────────────────────────────
    navy_fill  = PatternFill("solid", fgColor=HEX_NAVY)
    blue_fill  = PatternFill("solid", fgColor=HEX_BLUE)
    gold_fill  = PatternFill("solid", fgColor=HEX_GOLD)
    cyan_fill  = PatternFill("solid", fgColor=HEX_CYAN)
    light_fill = PatternFill("solid", fgColor=HEX_LIGHT)
    stripe_fill= PatternFill("solid", fgColor=HEX_STRIPE)
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    thin_blue  = Side(style="thin",   color=HEX_BORDER)
    thick_navy = Side(style="medium", color=HEX_NAVY)
    no_border  = Side(style=None)
    hdr_border = Border(bottom=thick_navy, top=Side(style="thin", color=HEX_NAVY),
                        left=thin_blue, right=thin_blue)

    n_cols = max(len(headers), 1)
    last_col = get_column_letter(n_cols)

    # ── Row 1: full-width navy brand banner ────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"  {BRAND_NAME}"
    ws["A1"].font      = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    ws["A1"].fill      = navy_fill
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Row 2: tagline ─────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"  {BRAND_TAGLINE}"
    ws["A2"].font      = Font(color="DBEAFE", size=8, italic=True, name="Calibri")
    ws["A2"].fill      = blue_fill
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # ── Row 3: contact line ────────────────────────────────────────────────
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"  {BRAND_ADDRESS}   |   {BRAND_WEBSITE}   |   {BRAND_EMAIL}   |   {BRAND_PHONE}"
    ws["A3"].font      = Font(color="DBEAFE", size=7, name="Calibri")
    ws["A3"].fill      = blue_fill
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 14

    # ── Row 4: gold separator ──────────────────────────────────────────────
    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"].fill      = gold_fill
    ws.row_dimensions[4].height = 4

    # ── Row 5: cyan accent bar ─────────────────────────────────────────────
    ws.merge_cells(f"A5:{last_col}5")
    ws["A5"].fill      = cyan_fill
    ws.row_dimensions[5].height = 2

    # ── Row 6: report title + generated timestamp ──────────────────────────
    ws.row_dimensions[6].height = 22
    ws.merge_cells(f"A6:{last_col}6")
    report_title = filename.replace("-", " ").replace("_", " ").title()
    ws["A6"] = f"  {report_title}"
    ws["A6"].font      = Font(color=HEX_NAVY, bold=True, size=12, name="Calibri")
    ws["A6"].fill      = light_fill
    ws["A6"].alignment = Alignment(horizontal="left", vertical="center")

    # Generated timestamp in last col (write after merge)
    ws.row_dimensions[7].height = 14
    ws.merge_cells(f"A7:{last_col}7")
    # Show subtitle (e.g. registration counts) if provided, otherwise timestamp
    row7_text = subtitle if subtitle else f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}"
    ws["A7"] = f"  {row7_text}"
    ws["A7"].font      = Font(color=HEX_MID, size=8, italic=True, name="Calibri")
    ws["A7"].fill      = light_fill
    ws["A7"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Row 8: spacer ──────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 6

    # ── Row 9: column headers ──────────────────────────────────────────────
    HDR_ROW = 9
    ws.row_dimensions[HDR_ROW].height = 22
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=HDR_ROW, column=col_idx, value=str(h).upper())
        cell.font      = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
        cell.fill      = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = hdr_border

    # ── Data rows ──────────────────────────────────────────────────────────
    for r_idx, row in enumerate(rows):
        excel_row  = HDR_ROW + 1 + r_idx
        is_even    = r_idx % 2 == 0
        row_fill   = white_fill if is_even else stripe_fill
        ws.row_dimensions[excel_row].height = 16

        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(
                row=excel_row, column=c_idx,
                value="" if val is None else str(val),
            )
            cell.fill      = row_fill
            cell.font      = Font(size=9, name="Calibri", color=HEX_DARK)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = Border(
                bottom=Side(style="thin", color="E5E7EB"),
                left=Side(style="thin",   color="E5E7EB"),
                right=Side(style="thin",  color="E5E7EB"),
            )

    # ── Column widths ──────────────────────────────────────────────────────
    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1])) + 2
        for cell in ws[col_letter]:
            if cell.value and not getattr(cell, 'data_type', None) == 'n':
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # ── Freeze panes below header row ─────────────────────────────────────
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # ── Save ───────────────────────────────────────────────────────────────
    stream = BytesIO()
    wb.save(stream)
    resp = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return resp


# ══════════════════════════════════════════════════════════════════════════════
# PDF ENGINE
# ══════════════════════════════════════════════════════════════════════════════

def pdf_response(filename, title, headers, rows, subtitle=""):
    if rl_canvas is None:
        text = title + "\n" + "\n".join(
            [" | ".join(str(h) for h in headers)] +
            [" | ".join("—" if (v is None or v == "") else str(v) for v in row) for row in rows]
        )
        resp = HttpResponse(text, content_type="text/plain")
        resp["Content-Disposition"] = f'attachment; filename="{filename}.txt"'
        return resp

    stream  = BytesIO()
    n_cols  = max(len(headers), 1)
    pw, ph  = (landscape(A4) if n_cols > 5 else A4)
    usable  = pw - 36 * mm

    doc = SimpleDocTemplate(
        stream,
        pagesize=(pw, ph),
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=22 * mm,  bottomMargin=24 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()

    # ── Custom paragraph styles ────────────────────────────────────────────
    def ps(name, **kw):
        base = kw.pop("parent", styles["Normal"])
        return ParagraphStyle(name, parent=base, **kw)

    st_brand = ps("Brand",
        fontSize=15, fontName="Helvetica-Bold",
        textColor=C_WHITE, leading=18,
    )
    st_tag = ps("Tag",
        fontSize=8, fontName="Helvetica",
        textColor=colors.HexColor("#DBEAFE"), leading=11,
    )
    st_contact = ps("Contact",
        fontSize=6.5, fontName="Helvetica",
        textColor=colors.HexColor("#BFDBFE"), leading=9,
    )
    st_report_title = ps("ReportTitle",
        fontSize=12, fontName="Helvetica-Bold",
        textColor=C_WHITE, leading=15, spaceBefore=4,
    )
    st_gen = ps("Gen",
        fontSize=7, fontName="Helvetica",
        textColor=colors.HexColor("#E5E7EB"), leading=9,
    )
    st_hdr = ps("Hdr",
        fontSize=8, fontName="Helvetica-Bold",
        textColor=C_WHITE, leading=10, wordWrap="CJK",
        alignment=TA_CENTER,
    )
    st_cell = ps("Cell",
        fontSize=8, fontName="Helvetica",
        textColor=C_DARK, leading=10, wordWrap="CJK",
    )
    st_cell_center = ps("CellC",
        fontSize=8, fontName="Helvetica",
        textColor=C_DARK, leading=10, wordWrap="CJK",
        alignment=TA_CENTER,
    )

    # ── Brand header block ─────────────────────────────────────────────────
    logo_path = _logo_path()
    brand_content = [
        Paragraph(BRAND_NAME, st_brand),
        Paragraph(BRAND_TAGLINE, st_tag),
        Spacer(1, 2),
        Paragraph(f"{BRAND_ADDRESS}  ·  {BRAND_WEBSITE}", st_contact),
        Paragraph(f"{BRAND_EMAIL}  ·  {BRAND_PHONE}", st_contact),
        Spacer(1, 5),
        Paragraph(title, st_report_title),
        Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}", st_gen),
    ]

    if logo_path:
        from reportlab.platypus import Image as RLImage
        try:
            logo_img = RLImage(logo_path, width=52, height=52)
            hdr_data  = [[logo_img, brand_content]]
            hdr_widths = [60, usable - 60]
        except Exception:
            hdr_data  = [[brand_content]]
            hdr_widths = [usable]
    else:
        hdr_data  = [[brand_content]]
        hdr_widths = [usable]

    hdr_table = Table(hdr_data, colWidths=hdr_widths)
    hdr_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), C_NAVY),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
        ("TOPPADDING",    (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))

    # Gold separator
    gold_bar = Table([[""]], colWidths=[usable], rowHeights=[4])
    gold_bar.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), C_GOLD),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # Cyan thin separator
    cyan_bar = Table([[""]], colWidths=[usable], rowHeights=[2])
    cyan_bar.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), C_CYAN),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))

    # ── Summary strip (row count) ──────────────────────────────────────────
    summary_ps = ps("Summary",
        fontSize=8, fontName="Helvetica",
        textColor=C_MID, leading=10,
    )
    summary_text = (
        subtitle if subtitle
        else (
            f"<b>{len(rows)}</b> record{'s' if len(rows) != 1 else ''}  ·  "
            f"Report period: {datetime.now().strftime('%B %Y')}"
        )
    )
    summary_para = Paragraph(summary_text, summary_ps)
    summary_table = Table([[summary_para]], colWidths=[usable])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), C_LIGHT),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.5, colors.HexColor("#BFDBFE")),
    ]))

    # ── Data table ─────────────────────────────────────────────────────────
    def fmt(v):
        return str(v) if v is not None and str(v).strip() not in ("", "—") else "—"

    # Smart column alignment: numeric/date-like → center, else left
    def is_numeric_col(col_idx):
        vals = [row[col_idx] for row in rows if col_idx < len(row)]
        hits = 0
        for v in vals[:20]:
            s = str(v).strip().replace("h", "").replace(",", "")
            try:
                float(s)
                hits += 1
            except ValueError:
                pass
        return hits > len(vals[:20]) * 0.6 if vals else False

    # Compute smart col widths based on content
    col_weights = []
    for ci in range(n_cols):
        max_w = len(str(headers[ci]))
        for row in rows[:50]:
            if ci < len(row):
                max_w = max(max_w, min(len(str(row[ci])), 40))
        col_weights.append(max(max_w, 6))
    total_weight = sum(col_weights)
    col_widths = [usable * w / total_weight for w in col_weights]

    tbl_rows = [[Paragraph(str(h), st_hdr) for h in headers]]
    for row in rows:
        tbl_rows.append([
            Paragraph(fmt(row[ci]) if ci < len(row) else "—",
                      st_cell_center if is_numeric_col(ci) else st_cell)
            for ci in range(n_cols)
        ])

    data_tbl = Table(tbl_rows, colWidths=col_widths, repeatRows=1)

    row_cmds = [
        # Header
        ("BACKGROUND",    (0, 0), (-1, 0),  C_NAVY),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  C_WHITE),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, 0),  7),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  7),
        # Body cells
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("TOPPADDING",    (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        # Outer border — navy
        ("BOX",           (0, 0), (-1, -1), 1,   C_BLUE),
        # Inner grid — light
        ("INNERGRID",     (0, 1), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        # Header bottom separator
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, C_GOLD),
        # Alternating rows
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [C_WHITE, C_LIGHT]),
    ]
    data_tbl.setStyle(TableStyle(row_cmds))

    # ── Footer callback ────────────────────────────────────────────────────
    def on_page(canvas_obj, doc_obj):
        canvas_obj.saveState()
        y = 16 * mm

        # Footer rule
        canvas_obj.setStrokeColor(C_GOLD)
        canvas_obj.setLineWidth(0.8)
        canvas_obj.line(18 * mm, y, pw - 18 * mm, y)

        # Left: copyright
        canvas_obj.setFillColor(C_MID)
        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.drawString(
            18 * mm, y - 5,
            f"© {datetime.now().year} {BRAND_NAME}   ·   {BRAND_WEBSITE}   ·   {BRAND_EMAIL}",
        )
        # Right: page number
        canvas_obj.setFillColor(C_NAVY)
        canvas_obj.setFont("Helvetica-Bold", 7)
        canvas_obj.drawRightString(
            pw - 18 * mm, y - 5,
            f"Page {doc_obj.page}",
        )
        canvas_obj.restoreState()

    # ── Assemble ───────────────────────────────────────────────────────────
    elements = [
        hdr_table,
        gold_bar,
        cyan_bar,
        Spacer(1, 3 * mm),
        summary_table,
        Spacer(1, 4 * mm),
        data_tbl,
    ]

    doc.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    resp = HttpResponse(stream.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return resp
